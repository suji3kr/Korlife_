
from bs4 import BeautifulSoup
import requests
import urllib.request as req
import os
import openpyxl
import datetime
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Color, Border, Side

#기존 엑셀 파일 삭제
if os.path.exists("./지니_크롤링.xlsx"):
    os.remove("./지니_크롤링.xlsx")

code = requests.get("https://www.genie.co.kr/chart/top200", headers={"User-Agent":"Mozilla/5.0"})
soup = BeautifulSoup(code.text, "html.parser")
title = soup.select("a.title.ellipsis")
name = soup.select("a.artist.ellipsis")
album = soup.select("a.albumtitle.ellipsis")
img = soup.select("a.cover > img")


# 이미지 저장할 폴더 생성
if not os.path.exists("./지니이미지"):
    os.mkdir("./지니이미지")

#엑셀 파일 생성
if not os.path.exists("./지니_크롤링.xlsx"):
    openpyxl.Workbook().save("./지니_크롤링.xlsx")

#엑셀 파일 불러오기
book = openpyxl.load_workbook("./지니_크롤링.xlsx")

#쓸데 없는 시트 지우기
if "Sheet" in book.sheetnames:
    book.remove(book["Sheet"])
sheet=book.create_sheet()
now = datetime.datetime.now()
sheet.title = f"{now.year}년 {now.month}월 {now.day}일 {now.hour}시 {now.minute}분 {now.second}초"

#열 너비 조절
sheet.cell(row=1, column=1).value="앨범 이미지"
sheet.cell(row=1, column=2).value="곡 이름"
sheet.cell(row=1, column=3).value="가수 이름"
sheet.cell(row=1, column=4).value="앨범 이름"
sheet.column_dimensions["A"].width = 15
sheet.column_dimensions["B"].width = 50
sheet.column_dimensions["C"].width = 30
sheet.column_dimensions["D"].width = 50

    
# img는 HTML 요소 리스트여야 한다.
 # img_tag에서 src 속성 가져오기
  
for i in range(len(title)):  
    img_url=img[i].attrs["src"]
    if img_url.startswith("//"):
       img_url = "https:" + img_url
    img_file_name = f"./지니이미지/{i+1}.png"
    req.urlretrieve(img_url, img_file_name)   
    print(f"{i+1}위. {title[i].get_text(strip=True)} - {name[i].get_text(strip=True)}")
    img_for_excel = Image(img_file_name)
    sheet.add_image(img_for_excel, f"A{i+2}")
    sheet.cell(row = i + 2, column = 2).value = title[i].get_text(strip=True)
    sheet.cell(row = i + 2 , column = 3).value = name[i].get_text(strip=True)
    sheet.cell(row = i + 2 , column = 4).value = album[i].get_text(strip=True)
    sheet.row_dimensions[i+2].height = 90

#폰트
bold_font = Font(size=12, bold=True)
font = Font(size=12)

#가운데 정렬
alignment_center = Alignment(horizontal="center", vertical="center")

#셀 배경색
color_sky = PatternFill(patternType="solid", fgColor=Color("ADD8E6"))

#테두리
border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

#첫 행에 서식 지정
for row in sheet["A1:D1"]:
    for cell in row:
        cell.font = bold_font
        cell.alignment = alignment_center
        cell.fill = color_sky
        cell.border = border

#데이터 부분에 서식 지정
for row in sheet["A2:D101"]:
    for cell in row:
        cell.font = font
        cell.alignment = alignment_center
        cell.border = border

book.save("./지니_크롤링.xlsx")


#